from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook

from imports import create_import_template, parse_import_workbook
from logging_setup import setup_runner_logging
from order_builders import build_order_spec, template_requires_limit, template_side
from pricing import price_preview
from schemas import (
    AccountSnapshot,
    BrokerOrder,
    ExecutionProfile,
    ImportOrderRow,
    PositionSnapshot,
    PreviewRow,
    Side,
    SortPreset,
)
from schwab_client import DEFAULT_CALLBACK_TIMEOUT_SECONDS, SchwabClient, SchwabClientError
from storage import db_connection, ensure_runtime_dirs, get_app_paths, init_db, save_setting


logger = setup_runner_logging()


def now_iso() -> str:
    return datetime.now().astimezone().isoformat()


def _load_profile(args: dict[str, Any]) -> ExecutionProfile:
    raw = args.get("execution_profile") or {}
    profile = ExecutionProfile.model_validate(raw)
    save_setting("execution_profile", profile.model_dump(mode="json"))
    return profile


def _side_for_quantity(quantity: float) -> Side:
    return Side.BUY if quantity > 0 else Side.SELL


def _reference_notional(row: PreviewRow, quantity: float, unit_price: float | None) -> float | None:
    if unit_price is None:
        return None
    return abs(quantity) * unit_price


def _strict_market_data(task_name: str) -> bool:
    return task_name in {"refresh_quotes", "place_orders"}


def _build_preview_rows(
    import_rows: list[ImportOrderRow],
    profile: ExecutionProfile,
    quote_map: dict[str, Any] | None = None,
    task_name: str = "validate_import",
) -> list[PreviewRow]:
    preview_rows: list[PreviewRow] = []
    strict_market_data = _strict_market_data(task_name)
    limit_required = template_requires_limit(profile.order_template)
    expected_side = template_side(profile.order_template)

    for import_row in import_rows:
        side = _side_for_quantity(import_row.quantity)
        quote = (quote_map or {}).get(import_row.symbol)
        preview = PreviewRow(
            row_number=import_row.row_number,
            account_number=import_row.account_number,
            symbol=import_row.symbol,
            quantity=import_row.quantity,
            side=side,
            enabled=import_row.enabled,
            note=import_row.note,
            limit_price_override=import_row.limit_price_override,
            take_profit_price=import_row.take_profit_price,
            stop_price=import_row.stop_price,
        )

        if side != expected_side:
            preview.validation_errors.append(
                f"Selected order template {profile.order_template.value} requires {expected_side.value} rows."
            )

        if strict_market_data and quote is None:
            preview.validation_errors.append("No quote was returned for this symbol.")

        if quote is not None:
            preview.open_price = quote.open_price
            preview.high_price = quote.high_price
            preview.low_price = quote.low_price
            preview.close_price = quote.close_price
            preview.last_price = quote.last_price
            preview.bid_price = quote.bid_price
            preview.ask_price = quote.ask_price

        pricing, pricing_errors = price_preview(
            side=side,
            last_price=preview.last_price,
            bid_price=preview.bid_price,
            ask_price=preview.ask_price,
            limit_price_override=import_row.limit_price_override,
            profile=profile,
            limit_required=limit_required,
            strict_market_data=strict_market_data,
        )
        preview.nbbo_spread = pricing["nbbo_spread"]
        preview.midpoint = pricing["midpoint"]
        preview.spread_percent = pricing["spread_percent"]
        preview.delta = pricing["delta"]
        preview.buy_limit = pricing["buy_limit"]
        preview.sell_limit = pricing["sell_limit"]
        preview.chosen_limit_price = pricing["chosen_limit_price"]
        preview.estimated_notional = _reference_notional(
            preview,
            import_row.quantity,
            pricing["estimated_notional_unit_price"],
        )
        preview.validation_errors.extend(pricing_errors)

        can_build_payload = not (
            not strict_market_data
            and limit_required
            and preview.chosen_limit_price is None
            and import_row.limit_price_override is None
        )

        if import_row.enabled and not preview.validation_errors and can_build_payload:
            try:
                preview.order_payload = build_order_spec(
                    order_template=profile.order_template,
                    account_number=import_row.account_number,
                    symbol=import_row.symbol,
                    quantity=import_row.quantity,
                    chosen_limit_price=preview.chosen_limit_price,
                    take_profit_price=import_row.take_profit_price,
                    stop_price=import_row.stop_price,
                    profile=profile,
                )
            except Exception as exc:  # noqa: BLE001
                preview.validation_errors.append(str(exc))

        if preview.validation_errors:
            preview.local_status = "INVALID"
        elif not import_row.enabled:
            preview.local_status = "DISABLED"
        elif not strict_market_data and limit_required and preview.chosen_limit_price is None:
            preview.local_status = "NEEDS_QUOTES"
            preview.local_status_detail = "Refresh quotes to compute limit pricing."
        else:
            preview.local_status = "READY"
        preview_rows.append(preview)

    return _apply_execution_order(preview_rows, profile.sort_preset)


def _apply_execution_order(preview_rows: list[PreviewRow], sort_preset: SortPreset) -> list[PreviewRow]:
    ready_rows = [row for row in preview_rows if row.enabled and not row.validation_errors]
    other_rows = [row for row in preview_rows if row not in ready_rows]

    def notional(row: PreviewRow) -> float:
        return float(row.estimated_notional or 0.0)

    if sort_preset == SortPreset.FILE_ORDER:
        ordered = sorted(ready_rows, key=lambda row: row.row_number)
    elif sort_preset == SortPreset.ABS_NOTIONAL_DESC:
        ordered = sorted(ready_rows, key=notional, reverse=True)
    elif sort_preset == SortPreset.SELLS_DESC_THEN_BUYS_DESC:
        sells = sorted((row for row in ready_rows if row.side == Side.SELL), key=notional, reverse=True)
        buys = sorted((row for row in ready_rows if row.side == Side.BUY), key=notional, reverse=True)
        ordered = [*sells, *buys]
    else:
        sells = sorted((row for row in ready_rows if row.side == Side.SELL), key=notional, reverse=True)
        buys = sorted((row for row in ready_rows if row.side == Side.BUY), key=notional)
        ordered = [*sells, *buys]

    for index, row in enumerate(ordered, start=1):
        row.execution_sequence = index
    for row in other_rows:
        row.execution_sequence = None
    return sorted(preview_rows, key=lambda row: (row.execution_sequence is None, row.execution_sequence or row.row_number, row.row_number))


def _persist_account_snapshots(snapshots: list[AccountSnapshot]) -> None:
    init_db()
    updated_at = now_iso()
    with db_connection() as conn:
        conn.execute("DELETE FROM account_snapshots")
        conn.executemany(
            """
            INSERT INTO account_snapshots(account_number, account_name, account_hash, cash_available, liquidation_value, raw_json, updated_at)
            VALUES(?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row.account_number,
                    row.account_name,
                    row.account_hash,
                    row.cash_available,
                    row.liquidation_value,
                    json.dumps(row.raw),
                    updated_at,
                )
                for row in snapshots
            ],
        )


def _persist_position_snapshots(snapshots: list[PositionSnapshot]) -> None:
    init_db()
    updated_at = now_iso()
    with db_connection() as conn:
        conn.execute("DELETE FROM position_snapshots")
        conn.executemany(
            """
            INSERT INTO position_snapshots(account_number, symbol, account_name, account_hash, average_price, quantity, value, day_pl, raw_json, updated_at)
            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row.account_number,
                    row.symbol,
                    row.account_name,
                    row.account_hash,
                    row.average_price,
                    row.quantity,
                    row.value,
                    row.day_pl,
                    json.dumps(row.raw),
                    updated_at,
                )
                for row in snapshots
            ],
        )


def _persist_broker_orders(rows: list[BrokerOrder]) -> None:
    init_db()
    updated_at = now_iso()
    with db_connection() as conn:
        conn.execute("DELETE FROM broker_orders")
        conn.executemany(
            """
            INSERT INTO broker_orders(order_id, account_name, account_number, account_hash, quantity, symbol, price, entered_time, time_in_force, session, status, status_details, cost_basis_method, cancelable, raw_json, updated_at)
            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row.order_id,
                    row.account_name,
                    row.account_number,
                    row.account_hash,
                    row.quantity,
                    row.symbol,
                    row.price,
                    row.entered_time,
                    row.time_in_force,
                    row.session,
                    row.status,
                    row.status_details,
                    row.cost_basis_method,
                    1 if row.cancelable else 0,
                    json.dumps(row.raw),
                    updated_at,
                )
                for row in rows
            ],
        )


def _persist_placed_order(
    *,
    run_id: str,
    request_id: str,
    preview: PreviewRow,
    account_hash: str | None,
    http_status: int | None,
    location: str | None,
    response_body: str,
) -> None:
    init_db()
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO placed_orders(run_id, request_id, row_number, account_number, account_hash, symbol, quantity, order_id, http_status, location, response_body, local_status, local_status_detail, payload_json, created_at)
            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                request_id,
                preview.row_number,
                preview.account_number,
                account_hash,
                preview.symbol,
                preview.quantity,
                preview.order_id,
                http_status,
                location,
                response_body,
                preview.local_status,
                preview.local_status_detail,
                json.dumps(preview.order_payload or {}),
                now_iso(),
            ),
        )


def _extract_order_id(location: str | None) -> str | None:
    if not location:
        return None
    tail = location.rstrip("/").split("/")[-1]
    return tail.split("?")[0] or None


def task_login(args: dict[str, Any]) -> dict[str, Any]:
    client = SchwabClient()
    app_key = str(args.get("app_key", "")).strip()
    app_secret = str(args.get("app_secret", "")).strip()
    callback_url = str(args.get("callback_url", "")).strip()
    if app_key and app_secret and callback_url:
        client.save_credentials(app_key, app_secret, callback_url)
    else:
        client.load_credentials()
    session = client.connect(
        force_login=bool(args.get("force_login")),
        interactive=False,
        callback_timeout=DEFAULT_CALLBACK_TIMEOUT_SECONDS,
    )
    linked = client.get_account_numbers()
    return {
        "login_status": client.login_status(),
        "linked_account_count": len(linked),
        "api_key_suffix": client.load_credentials().get("app_key", "")[-4:],
        "session_class": type(session).__name__,
    }


def task_validate_import(args: dict[str, Any]) -> dict[str, Any]:
    profile = _load_profile(args)
    import_path = Path(args["import_path"]).resolve()
    rows, errors = parse_import_workbook(import_path)
    preview_rows = _build_preview_rows(rows, profile, quote_map=None, task_name="validate_import")
    return {
        "import_path": str(import_path),
        "errors": [error.model_dump(mode="json") for error in errors],
        "preview_rows": [row.model_dump(mode="json") for row in preview_rows],
    }


def task_refresh_quotes(args: dict[str, Any]) -> dict[str, Any]:
    profile = _load_profile(args)
    import_path = Path(args["import_path"]).resolve()
    rows, errors = parse_import_workbook(import_path)
    if errors:
        return {
            "import_path": str(import_path),
            "errors": [error.model_dump(mode="json") for error in errors],
            "preview_rows": [],
        }
    client = SchwabClient()
    quote_map = client.get_quotes([row.symbol for row in rows])
    preview_rows = _build_preview_rows(rows, profile, quote_map=quote_map, task_name="refresh_quotes")
    return {
        "import_path": str(import_path),
        "errors": [],
        "preview_rows": [row.model_dump(mode="json") for row in preview_rows],
    }


def task_place_orders(args: dict[str, Any]) -> dict[str, Any]:
    profile = _load_profile(args)
    import_path = Path(args["import_path"]).resolve()
    rows, errors = parse_import_workbook(import_path)
    if errors:
        return {
            "import_path": str(import_path),
            "errors": [error.model_dump(mode="json") for error in errors],
            "preview_rows": [],
        }

    client = SchwabClient()
    quote_map = client.get_quotes([row.symbol for row in rows])
    preview_rows = _build_preview_rows(rows, profile, quote_map=quote_map, task_name="place_orders")
    run_id = args.get("request_id") or f"run-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    init_db()
    with db_connection() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO order_runs(run_id, created_at, import_path, execution_profile_json, summary_json) VALUES(?, ?, ?, ?, ?)",
            (run_id, now_iso(), str(import_path), json.dumps(profile.model_dump(mode="json")), None),
        )

    hash_map = client.get_account_hash_map()
    submitted = 0
    rejected = 0
    for preview in preview_rows:
        if not preview.enabled or preview.validation_errors or preview.order_payload is None:
            continue
        account_hash = hash_map.get(preview.account_number)
        if not account_hash:
            preview.local_status = "REJECTED"
            preview.local_status_detail = "Account not linked."
            rejected += 1
            _persist_placed_order(
                run_id=run_id,
                request_id=run_id,
                preview=preview,
                account_hash=None,
                http_status=None,
                location=None,
                response_body="Account not linked",
            )
            continue

        if profile.preview_only:
            preview.local_status = "PREVIEW_ONLY"
            preview.local_status_detail = "Preview only mode enabled."
            continue

        response = client.place_order(account_hash, preview.order_payload)
        location = response.headers.get("Location")
        preview.order_id = _extract_order_id(location)
        if 200 <= response.status_code < 300:
            preview.local_status = "SUBMITTED"
            preview.local_status_detail = f"OrderId: {preview.order_id}" if preview.order_id else ""
            submitted += 1
        else:
            preview.local_status = "REJECTED"
            preview.local_status_detail = response.text[:500]
            rejected += 1

        _persist_placed_order(
            run_id=run_id,
            request_id=run_id,
            preview=preview,
            account_hash=account_hash,
            http_status=response.status_code,
            location=location,
            response_body=response.text,
        )

    return {
        "run_id": run_id,
        "import_path": str(import_path),
        "errors": [],
        "submitted_count": submitted,
        "rejected_count": rejected,
        "preview_rows": [row.model_dump(mode="json") for row in preview_rows],
    }


def task_refresh_accounts(args: dict[str, Any]) -> dict[str, Any]:
    client = SchwabClient()
    snapshots = client.get_accounts_snapshot()
    _persist_account_snapshots(snapshots)
    return {"accounts": [row.model_dump(mode="json") for row in snapshots]}


def task_refresh_orders(args: dict[str, Any]) -> dict[str, Any]:
    profile = _load_profile(args)
    client = SchwabClient()
    rows = client.get_orders_snapshot(profile.orders_lookback_days)
    _persist_broker_orders(rows)
    return {"orders": [row.model_dump(mode="json") for row in rows]}


def task_refresh_portfolio(args: dict[str, Any]) -> dict[str, Any]:
    client = SchwabClient()
    rows = client.get_positions_snapshot()
    _persist_position_snapshots(rows)
    return {"positions": [row.model_dump(mode="json") for row in rows]}


def _load_snapshot_rows(table_name: str) -> list[dict[str, Any]]:
    init_db()
    with db_connection() as conn:
        rows = conn.execute(f"SELECT * FROM {table_name}").fetchall()
    return [dict(row) for row in rows]


def task_export_snapshot(args: dict[str, Any]) -> dict[str, Any]:
    paths = ensure_runtime_dirs(get_app_paths())
    accounts = _load_snapshot_rows("account_snapshots")
    positions = _load_snapshot_rows("position_snapshots")
    orders = _load_snapshot_rows("broker_orders")

    if not accounts:
        accounts = task_refresh_accounts({})["accounts"]
    if not positions:
        positions = task_refresh_portfolio({})["positions"]
    if not orders:
        profile = _load_profile(args)
        orders = task_refresh_orders({"execution_profile": profile.model_dump(mode="json")})["orders"]

    workbook = Workbook()
    sheet_accounts = workbook.active
    sheet_accounts.title = "Accounts"
    sheet_portfolio = workbook.create_sheet("Portfolio")
    sheet_orders = workbook.create_sheet("Order Status")

    account_headers = ["ACCOUNT NAME", "ACCOUNT NUMBER", "ACCOUNT HASH", "CASH AVAILABLE", "LIQUIDATION VALUE"]
    portfolio_headers = ["ACCOUNT NAME", "ACCOUNT NUMBER", "SYMBOL", "AVG PRICE", "QTY", "VALUE", "DAY P/L"]
    order_headers = [
        "ACCOUNT NAME",
        "ACCOUNT NUMBER",
        "ORDER ID",
        "STATUS",
        "QTY",
        "SYMBOL",
        "PRICE",
        "ENTERED",
        "TIME IN FORCE",
        "SESSION",
        "COST BASIS METHOD",
        "STATUS DETAILS",
    ]

    sheet_accounts.append(account_headers)
    for row in sorted(accounts, key=lambda item: (item["account_name"], item["account_number"])):
        sheet_accounts.append(
            [row["account_name"], row["account_number"], row["account_hash"], row["cash_available"], row["liquidation_value"]]
        )

    sheet_portfolio.append(portfolio_headers)
    for row in sorted(positions, key=lambda item: (item["account_name"], item["symbol"])):
        sheet_portfolio.append(
            [
                row["account_name"],
                row["account_number"],
                row["symbol"],
                row["average_price"],
                row["quantity"],
                row["value"],
                row["day_pl"],
            ]
        )

    sheet_orders.append(order_headers)
    for row in sorted(orders, key=lambda item: item.get("entered_time") or "", reverse=True):
        sheet_orders.append(
            [
                row["account_name"],
                row["account_number"],
                row["order_id"],
                row["status"],
                row["quantity"],
                row["symbol"],
                row["price"],
                row["entered_time"],
                row["time_in_force"],
                row["session"],
                row["cost_basis_method"],
                row["status_details"],
            ]
        )

    output_path = paths.exports_dir / f"SchwabData_{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    workbook.save(output_path)
    return {
        "export_path": str(output_path),
        "accounts_count": len(accounts),
        "positions_count": len(positions),
        "orders_count": len(orders),
    }


def task_create_import_template(args: dict[str, Any]) -> dict[str, Any]:
    paths = ensure_runtime_dirs(get_app_paths())
    target = args.get("template_path") or str(paths.imports_dir / "order_template.xlsx")
    path = create_import_template(target)
    return {"template_path": str(path)}


TASKS: dict[str, Callable[[dict[str, Any]], dict[str, Any]]] = {
    "login": task_login,
    "validate_import": task_validate_import,
    "refresh_quotes": task_refresh_quotes,
    "place_orders": task_place_orders,
    "refresh_accounts": task_refresh_accounts,
    "refresh_orders": task_refresh_orders,
    "refresh_portfolio": task_refresh_portfolio,
    "export_snapshot": task_export_snapshot,
    "create_import_template": task_create_import_template,
}
