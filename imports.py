from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from schemas import ImportOrderRow, ImportValidationError


REQUIRED_HEADERS = ("account_number", "symbol", "quantity")
OPTIONAL_HEADERS = ("enabled", "limit_price_override", "take_profit_price", "stop_price", "note")
ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _normalize_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if text in {"", "1", "true", "yes", "y"}:
        return True
    if text in {"0", "false", "no", "n"}:
        return False
    raise ValueError(f"Unsupported enabled value: {value!r}")


def parse_import_workbook(path: str | Path) -> tuple[list[ImportOrderRow], list[ImportValidationError]]:
    workbook_path = Path(path)
    errors: list[ImportValidationError] = []
    rows: list[ImportOrderRow] = []

    if not workbook_path.exists():
        return [], [ImportValidationError(field_name="file", message=f"Import file not found: {workbook_path}")]

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    headers = [_normalize_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {header: index for index, header in enumerate(headers) if header}

    for required in REQUIRED_HEADERS:
        if required not in header_map:
            errors.append(
                ImportValidationError(
                    field_name=required,
                    message=f"Missing required header: {required}",
                )
            )

    if errors:
        return [], errors

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in row):
            continue

        payload: dict[str, Any] = {"row_number": row_number}
        for header in ALL_HEADERS:
            if header not in header_map:
                continue
            value = row[header_map[header]]
            payload[header] = value

        try:
            payload["enabled"] = _normalize_bool(payload.get("enabled", True))
            parsed = ImportOrderRow.model_validate(payload)
            if not parsed.account_number:
                raise ValueError("account_number is blank")
            if not parsed.symbol:
                raise ValueError("symbol is blank")
            if parsed.quantity == 0:
                raise ValueError("quantity cannot be zero")
            rows.append(parsed)
        except Exception as exc:  # noqa: BLE001
            errors.append(
                ImportValidationError(
                    row_number=row_number,
                    field_name="row",
                    message=str(exc),
                )
            )

    return rows, errors


def create_import_template(path: str | Path) -> Path:
    target = Path(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    for column, header in enumerate(ALL_HEADERS, start=1):
        sheet.cell(row=1, column=column, value=header)
    sheet.cell(row=2, column=1, value="12345678")
    sheet.cell(row=2, column=2, value="SPY")
    sheet.cell(row=2, column=3, value=10)
    sheet.cell(row=2, column=4, value=True)
    sheet.cell(row=2, column=8, value="Example row")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target

